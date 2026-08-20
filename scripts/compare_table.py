

import sys
import yaml
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.colors import to_rgba


root_dir = Path(__file__).resolve().parents[1]
if str(root_dir) not in sys.path:
    sys.path.insert(0, str(root_dir))

from src.env import Env


ESTIMATOR_LABELS = {
    "GJRGARCH":    "GJR-GARCH",
    "GARCH":       "GARCH",
    "EWMA":        "EWMA",
    "RealizedVol": "Realized Vol",
    "AR1":         "AR1",
    "BuyAndHold":  "Buy & Hold",
}

CONTROLLER_LABELS = {
    "NaiveScaling":           "Naïve",
    "VarianceScaling":        "Variance Scaled",
    "RegimeSwitchController": "Regime Switch",
    "ConstantWeight":         "Constant Weight",
}

CONTROLLER_ORDER = ["Naïve", "Variance Scaled", "Regime Switch", "Constant Weight"]
ESTIMATOR_ORDER  = ["GJR-GARCH", "GARCH", "EWMA", "Realized Vol", "AR1", "Buy & Hold"]
METRICS          = ["Sharpe", "Annual Vol", "Annual Return", "Max DD"]


DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5FAC"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"

METRIC_COLORS = {
    "Sharpe":        "1F497D",
    "Annual Vol":    "375623",
    "Annual Return": "7F3F00",
    "Max DD":        "632523",
}


# ── helpers ─────────────────

def _hex(h: str):
    """Convert 6-char hex string to matplotlib rgba tuple."""
    h = h.lstrip("#")
    return to_rgba(f"#{h}")

def _class_label(class_path: str, mapping: dict) -> str:
    class_name = class_path.rsplit(".", 1)[-1]
    return mapping.get(class_name, class_name)

def load_configs() -> list[dict]:
    strategies_dir = Env.path("strategies")
    configs = []
    for p in strategies_dir.glob("*.yaml"):
        try:
            with open(p, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f)
            if cfg:
                cfg["_yaml_path"] = p
                configs.append(cfg)
        except Exception as e:
            print(f"Could not read {p.name}: {e}")
    return configs

def compute_metrics(df: pd.DataFrame) -> dict:
    rets    = df["returns"].fillna(0.0)
    ann_ret = float(rets.mean() * 252)
    ann_vol = float(rets.std() * np.sqrt(252))
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else np.nan

    equity = df.get("equity_curve", None)
    if equity is None:
        equity = pd.Series(1000.0 * np.exp(np.cumsum(rets.values)), index=df.index)

    dd    = equity / equity.cummax() - 1
    max_dd = float(dd.min())
    ulcer  = float(np.sqrt(np.mean(dd.values ** 2)))

    return {
        "Sharpe": round(sharpe,  3),
        "Annual Vol": round(ann_vol, 4),
        "Annual Return": round(ann_ret, 4),
        "Max DD": round(max_dd,  4),
        "Ulcer Index": round(ulcer,   4),
    }

def build_data_table():
    results_dir = Env.path("results")
    configs = load_configs()
    rows = {}

    for cfg in configs:
        strat_name = cfg.get("name", "")
        est_path = cfg.get("estimator", {}).get("class", "")
        ctrl_path = cfg.get("controller", {}).get("class", "")
        est_label = _class_label(est_path, ESTIMATOR_LABELS)
        ctrl_label = _class_label(ctrl_path, CONTROLLER_LABELS)

        result_path = results_dir / f"{strat_name}.parquet"
        if not result_path.exists():
            print(f" No result for '{strat_name}' — run backtest first.")
            continue

        try:
            df = pd.read_parquet(result_path)
            metrics = compute_metrics(df)
        except Exception as e:
            print(f"Could not compute metrics for '{strat_name}': {e}")
            continue

        rows.setdefault(est_label, {})
        for metric, value in metrics.items():
            rows[est_label][(metric, ctrl_label)] = value

    if not rows:
        raise ValueError("No results found. Run run_backtests.py first.")

    all_controllers = sorted(
        {ctrl for vals in rows.values() for _, ctrl in vals.keys()},
        key=lambda c: CONTROLLER_ORDER.index(c) if c in CONTROLLER_ORDER else 999,
    )
    estimators = sorted(
        rows.keys(),
        key=lambda e: ESTIMATOR_ORDER.index(e) if e in ESTIMATOR_ORDER else 999,
    )

    multi_cols = pd.MultiIndex.from_product([METRICS, all_controllers])
    table = pd.DataFrame(index=estimators, columns=multi_cols, dtype=object)
    for est, vals in rows.items():
        for (metric, ctrl), value in vals.items():
            if (metric, ctrl) in table.columns:
                table.loc[est, (metric, ctrl)] = value

    return table, all_controllers


#Excel writer 

def _side(style="thin"):
    return Side(border_style=style, color="000000")

def _border(top=None, bottom=None, left=None, right=None):
    return Border(top=top, bottom=bottom, left=left, right=right)

def write_excel(table: pd.DataFrame, controllers: list[str], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Controller Comparison"

    n_ctrl = len(controllers)
    n_met  = len(METRICS)
    n_est  = len(table.index)

    ws.column_dimensions["A"].width = 18
    for col_i in range(2, 2 + n_met * n_ctrl):
        ws.column_dimensions[get_column_letter(col_i)].width = 15

    # title
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_met * n_ctrl)
    tc = ws.cell(row=1, column=1, value="Controller Comparison — Key Performance Metrics")
    tc.font = Font(bold=True, size=14, color=WHITE, name="Arial")
    tc.fill = PatternFill("solid", fgColor=DARK_BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    # metric group headers
    ws.row_dimensions[2].height = 22
    corner = ws.cell(row=2, column=1, value="Estimator")
    corner.font = Font(bold=True, color=WHITE, name="Arial")
    corner.fill = PatternFill("solid", fgColor=DARK_BLUE)
    corner.alignment = Alignment(horizontal="center", vertical="center")

    for m_i, metric in enumerate(METRICS):
        col_start = 2 + m_i * n_ctrl
        col_end = col_start + n_ctrl - 1
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        cell = ws.cell(row=2, column=col_start, value=metric)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor=METRIC_COLORS.get(metric, MID_BLUE))
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # controller sub-headers
    ws.row_dimensions[3].height = 18
    ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    for m_i in range(n_met):
        for c_i, ctrl in enumerate(controllers):
            col = 2 + m_i * n_ctrl + c_i
            cell = ws.cell(row=3, column=col, value=ctrl)
            cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=MID_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for r_i, estimator in enumerate(table.index):
        row = 4 + r_i
        row_fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_i % 2 == 0 else WHITE)
        ws.row_dimensions[row].height = 16

        ec = ws.cell(row=row, column=1, value=estimator)
        ec.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        ec.fill = PatternFill("solid", fgColor=DARK_BLUE)
        ec.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for m_i, metric in enumerate(METRICS):
            for c_i, ctrl in enumerate(controllers):
                col = 2 + m_i * n_ctrl + c_i
                value = table.loc[estimator, (metric, ctrl)]
                cell = ws.cell(row=row, column=col)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=10)

                if pd.isna(value) or value is None:
                    cell.value = "—"
                    cell.font = Font(name="Arial", size=10, color="999999")
                    continue

                val = float(value)
                if metric == "Sharpe":
                    cell.value = round(val, 3)
                    cell.number_format = "0.000"
                    if val >= 0.8:
                        cell.font = Font(name="Arial", size=10, color="375623", bold=True)
                    elif val < 0:
                        cell.font = Font(name="Arial", size=10, color="C00000")
                elif metric in ("Annual Vol", "Annual Return"):
                    cell.value = val
                    cell.number_format = "0.00%"
                    if metric == "Annual Return" and val < 0:
                        cell.font = Font(name="Arial", size=10, color="C00000")
                elif metric == "Max DD":
                    cell.value = val
                    cell.number_format = "0.00%"
                    if val < -0.20:
                        cell.font = Font(name="Arial", size=10, color="C00000")
                elif metric == "Ulcer Index":
                    cell.value = round(val, 4)
                    cell.number_format = "0.0000"

        for m_i in range(n_met - 1):
            sep_col = 2 + (m_i + 1) * n_ctrl - 1
            ws.cell(row=row, column=sep_col).border = _border(right=_side("medium"))

    thick = _side("medium")
    for col in range(1, 2 + n_met * n_ctrl):
        ws.cell(row=3, column=col).border = _border(bottom=thick)
        ws.cell(row=4 + n_est - 1, column=col).border = _border(bottom=thick)
        ws.cell(row=2, column=col).border = _border(top=thick)
    for row in range(2, 4 + n_est):
        ws.cell(row=row, column=1).border = _border(left=thick)
        ws.cell(row=row, column=1 + n_met * n_ctrl).border = _border(right=thick)

    ws.freeze_panes = "B4"
    wb.save(out_path)
    print(f"Excel saved : {out_path}")


# ── PNG writer

def _fmt_value(value, metric: str) -> str:
    """Format a cell value for display in the PNG table."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "—"
    val = float(value)
    if metric == "Sharpe":
        return f"{val:.3f}"
    elif metric in ("Annual Vol", "Annual Return", "Max DD"):
        return f"{val:.2%}"
    elif metric == "Ulcer Index":
        return f"{val:.4f}"
    return str(value)


def _cell_text_color(value, metric: str) -> str:
    """Return red for bad values, green for good Sharpe, else black."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "#999999"
    val = float(value)
    if metric == "Sharpe":
        if val >= 0.8:
            return "#1A5C1A"   # dark green
        if val < 0:
            return "#C00000"   # red
    if metric == "Annual Return" and val < 0:
        return "#C00000"
    if metric == "Max DD" and val < -0.20:
        return "#C00000"
    return "#000000"


def write_png(table: pd.DataFrame, controllers: list[str], out_path: Path):
    n_ctrl = len(controllers)
    n_met  = len(METRICS)
    n_est  = len(table.index)

    # Total columns: 1 (estimator) + n_met * n_ctrl
    total_cols = 1 + n_met * n_ctrl

    # Figure sizing
    col_w   = 1.6   # inches per data column
    est_w   = 1.8   # inches for estimator column
    row_h   = 0.38  # inches per row
    n_rows  = 3 + n_est  # title + metric header + ctrl header + data

    fig_w = est_w + n_met * n_ctrl * col_w
    fig_h = n_rows * row_h + 0.4   # small padding

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    # Helper: column x-positions
    col_x = [0]  # left edge of each column
    col_x.append(est_w)
    for _ in range(n_met * n_ctrl):
        col_x.append(col_x[-1] + col_w)
    col_x = col_x[:-1]  # trim extra

    col_widths = [est_w] + [col_w] * (n_met * n_ctrl)

    # Row y-positions (top-down)
    row_y = [fig_h - 0.2]  # title row top
    for i in range(n_rows):
        row_y.append(row_y[-1] - row_h)

    def draw_cell(row_idx, col_idx, text, bg_hex, text_color="#FFFFFF",
                  bold=False, fontsize=9, ha="center", colspan=1):
        x     = col_x[col_idx]
        y     = row_y[row_idx + 1]
        w     = sum(col_widths[col_idx:col_idx + colspan])
        h     = row_h
        rect  = mpatches.FancyBboxPatch(
            (x, y), w, h,
            boxstyle="square,pad=0",
            linewidth=0.5,
            edgecolor="#CCCCCC",
            facecolor=_hex(bg_hex),
        )
        ax.add_patch(rect)
        ax.text(
            x + w / 2 if ha == "center" else x + 0.08,
            y + h / 2,
            text,
            ha=ha, va="center",
            fontsize=fontsize,
            fontweight="bold" if bold else "normal",
            color=text_color,
            clip_on=True,
        )

    # ── Row 0: Title ──────────────────────────────────────────────────────────
    draw_cell(0, 0, "Controller Comparison — Key Performance Metrics",
              DARK_BLUE, bold=True, fontsize=11, colspan=total_cols)

    # ── Row 1: Metric group headers ───────────────────────────────────────────
    draw_cell(1, 0, "Estimator", DARK_BLUE, bold=True, fontsize=9)
    for m_i, metric in enumerate(METRICS):
        draw_cell(1, 1 + m_i * n_ctrl, metric,
                  METRIC_COLORS.get(metric, MID_BLUE),
                  bold=True, fontsize=9, colspan=n_ctrl)

    # ── Row 2: Controller sub-headers ─────────────────────────────────────────
    draw_cell(2, 0, "", DARK_BLUE)
    for m_i in range(n_met):
        for c_i, ctrl in enumerate(controllers):
            draw_cell(2, 1 + m_i * n_ctrl + c_i, ctrl,
                      MID_BLUE, bold=True, fontsize=8)

    # ── Rows 3+: Data ─────────────────────────────────────────────────────────
    for r_i, estimator in enumerate(table.index):
        row_bg = LIGHT_BLUE if r_i % 2 == 0 else WHITE
        draw_cell(3 + r_i, 0, estimator, DARK_BLUE,
                  bold=True, fontsize=9, ha="left")

        for m_i, metric in enumerate(METRICS):
            for c_i, ctrl in enumerate(controllers):
                value    = table.loc[estimator, (metric, ctrl)]
                txt      = _fmt_value(value, metric)
                txt_col  = _cell_text_color(value, metric)
                bold_val = (metric == "Sharpe" and value is not None
                            and not (isinstance(value, float) and np.isnan(value))
                            and float(value) >= 0.8)
                draw_cell(3 + r_i, 1 + m_i * n_ctrl + c_i,
                          txt, row_bg, text_color=txt_col,
                          bold=bold_val, fontsize=8.5)

    # Vertical separator lines between metric groups
    for m_i in range(1, n_met):
        sep_x = col_x[1 + m_i * n_ctrl]
        ax.plot([sep_x, sep_x],
                [row_y[n_rows], row_y[1]],
                color="#555555", linewidth=1.2, zorder=5)

    # Outer border
    outer = mpatches.FancyBboxPatch(
        (0, row_y[n_rows]), fig_w, fig_h - 0.2 - row_y[n_rows],
        boxstyle="square,pad=0",
        linewidth=1.5,
        edgecolor="#1F3864",
        facecolor="none",
        zorder=10,
    )
    ax.add_patch(outer)

    plt.tight_layout(pad=0)
    fig.savefig(out_path, dpi=200, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    print(f"✅  PNG saved   : {out_path}")


# ── main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("📊  Building controller comparison table...")

    table, controllers = build_data_table()

    print(f"\n  Estimators found : {list(table.index)}")
    print(f"  Controllers found: {controllers}")
    print(f"  Metrics          : {METRICS}\n")

    results_dir = Env.path("results")

    write_excel(table, controllers, results_dir / "controller_comparison.xlsx")
    write_png(  table, controllers, results_dir / "controller_comparison.png")
